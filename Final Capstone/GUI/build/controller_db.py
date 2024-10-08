import datetime
import pandas as pd
from collections import namedtuple
from dateutil.relativedelta import relativedelta
from main_util import set_notice
from openpyxl.reader.excel import load_workbook
from controller_model import get_best_prediction
import config


def refresh_database():
    global main_stats_month
    main_stats_month = pd.read_excel(config.DB_FILE, sheet_name='Month_Stats')
    return main_stats_month


def long_department_to_code(long_s: str):
    if long_s.upper() == 'ALL':
        return 'ALL'
    mapper = config.department_config[['Department Short Code', 'Department Long Code']].drop_duplicates()
    return mapper[mapper['Department Long Code'] == long_s]['Department Short Code'].iloc[0]


def code_to_long_department(code: str):
    if code.upper() == 'ALL':
        return 'ALL'
    mapper = config.department_config[['Department Short Code', 'Department Long Code']].drop_duplicates()
    return mapper[mapper['Department Short Code'] == code]['Department Long Code'].iloc[0]


def month2idx(months="Jul"):
    MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    return MONTHS.index(months) + 1


def get_training_months(predicted_year, predicted_month, department):
    TRAINING_WINDOW = 24
    TESTING_WINDOW = 3
    refresh_database()
    predict_year_month = datetime.date(day=1, month=predicted_month, year=predicted_year)
    start_date_train = predict_year_month - relativedelta(months=TRAINING_WINDOW + TESTING_WINDOW + 1)
    end_date_train = predict_year_month - relativedelta(months=1)
    filtered_df = main_stats_month[main_stats_month.Department == department]
    train_date_list = [x.strftime("%Y-%m") for x in pd.date_range(start_date_train, end_date_train)]
    train_df = filtered_df[filtered_df["Months"].isin(train_date_list)].sort_values('Months')
    return train_df


def create_month_year(year, month):
    return f"{int(year)}-{str(int(month)).zfill(2)}"


def get_all_months_data(year: int = 2024, month: int = 1, department='ALL'):
    def get_threshold(department):
        return eval(
            config.department_config[config.department_config['Department Short Code'] == department].Threshold.iloc[0])

    THRESHOLD = get_threshold(department)

    def get_colour(percentage):
        if percentage > THRESHOLD[0]:
            return 'red'
        elif percentage > THRESHOLD[1]:
            return 'orange'
        elif percentage > THRESHOLD[2]:
            return 'yellow'
        else:
            return 'green'

    department_data = main_stats_month[main_stats_month.Department == department]
    filtered = department_data[department_data.Months.str.contains(str(year))]
    MonthData = namedtuple(typename='MonthData', field_names=['month_txt', 'per_txt', 'rec_colour'])
    MONTH_DATA = {'Year': str(year)}
    MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    color = 'black'
    for i in range(1, 13):
        output = filtered.query(f"Months == '{create_month_year(year, i)}'")
        if len(output) == 0:

            percentage_str = "Not Avaliable"
            color = 'black'
        else:
            if output.fillna(0).iloc[0]['Demand'] != 0:
                percentage = output.iloc[0]['Actual_WTA']
                color = get_colour(percentage)
                percentage_str = f"{percentage:.1f}%"
            elif pd.isna(output.iloc[0]['Predicted_WTA']) == False:
                percentage = output.iloc[0]['Predicted_WTA']
                color = get_colour(percentage)
                percentage_str = f"({percentage:.1f}%)"
            else:
                percentage_str = "Not Avaliable"
                color = 'Black'

        #
        MONTH_DATA[MONTHS[i - 1]] = MonthData(MONTHS[i - 1], percentage_str, color)

    filtered_months = department_data[department_data.Months == f"{str(int(year))}-{str(int(month)).zfill(2)}"]
    if len(filtered_months) == 0:
        statistics = {f'{MONTHS[month - 1]} WTA': '-',
                      f'{MONTHS[month - 1]} Demand': '-',
                      f'{MONTHS[month - 1]} Supply': '-',
                      f'{MONTHS[month - 1]} Predicted WTA': '-',
                      f'{MONTHS[month - 1]} Total Expected': '-',
                      f'{MONTHS[month - 1]} Total Booked': '-',
                      }

    else:
        filtered_months = filtered_months.iloc[0].fillna(0)

        statistics = {f'{MONTHS[month - 1]} WTA': f'{filtered_months.Actual_WTA:.2f}%',
                      f'{MONTHS[month - 1]} Demand': f'{filtered_months.Demand:.0f}',
                      f'{MONTHS[month - 1]} Supply': f'{filtered_months.Supply:.0f}',
                      f'{MONTHS[month - 1]} Predicted WTA': f'{filtered_months.Predicted_WTA:.2f}%',
                      f'{MONTHS[month - 1]} Total Expected': f'{filtered_months.Predicted_Demand:.0f}',
                      f'{MONTHS[month - 1]} Total Booked': f'{filtered_months.Slot_booked:.0f}',
                      }

    return MONTH_DATA, statistics


def append_config_df(df):
    from openpyxl.reader.excel import load_workbook
    wb = load_workbook(config.CONFIGURATION_EXCEL)
    ws = wb.active
    temp = 2
    for idx, r in df.iterrows():
        ws.cell(temp, 1).value = r['Config Key']
        ws.cell(temp, 2).value = r['Config Value']
        temp += 1
        print(f"Added into {r['Config Key']}, {r['Config Value']}")
    wb.save(config.CONFIGURATION_EXCEL)
    wb.close()


def append_config():
    append_config_df(pd.DataFrame(config.CONFIGURATION))
    print("Completed Appending")


def regroup_specialty(dept):
    """
    Changing Subcodes to Short Codes
    :param dept:
    :return:
    """
    def get_mapper():
        mapper = config.department_config[['Department SubCodes', 'Department Short Code']].set_index(
            'Department SubCodes')
        mapper_dict = mapper['Department Short Code'].to_dict()
        return mapper_dict

    mapper = get_mapper()
    return mapper.get(dept, 'Unknown')


def get_supply(input_file_path, department_interested=['GENT', 'GE1H', 'GE1S', 'GOTO', 'GO1S', 'GGAS']):
    """
    Main Function in getting supply from SGH supply files
    :param input_file_path: Input File Path
    :param department_interested: Deparmtnent short codes that are interested
    :return:
    """
    COLS_TO_IDENTIFY = ['Resource', 'Resource Date', 'SessionType', 'Room', 'Slot Payment Class Type',
                        'Slot Visit Type']
    PRIORITY_LIST = {'Daily': 1, ',Weekly': 2, 'Week of Month': 3}
    INTERESTED_COLS = ['Resource Date', 'Dept of Resource', 'Resource', 'Rank of Resource', 'Specialty Code', 'Clinic',
                       'Room',
                       'Frequency',
                       'SessionType', 'Slot Payment Class Type', 'Slot Visit Type', 'Blocked Status', 'Total Booked',
                       'Total Slot Count']

    # Read the Excel file
    df = pd.read_excel(input_file_path)

    # 1. Replace column names with values from the first row (index 0)
    new_columns = df.iloc[0].values
    df.columns = new_columns

    # 2. Drop the first row
    df = df.drop(df.index[0])

    # 3. Reset the index
    df.reset_index(drop=True, inplace=True)

    # Create subset/Drop Columns
    df_subset = df[INTERESTED_COLS]
    # --- Apply filters ----#
    # Interest = New Case slots (NC)
    df_subset = df_subset[df_subset['Slot Visit Type'] == 'NC']
    # Interest = GEN + SUB slots
    df_subset = df_subset[df_subset['Slot Payment Class Type'] != 'PTE']
    # Remove Non-Doctor ranks
    df_subset = df_subset[~df_subset['Rank of Resource'].isin(['DO', 'DOS', 'NDR', ''])]
    # Specialty of interest based on dept of resource
    df_subset = df_subset[df_subset['Dept of Resource'].isin(department_interested)]
    # --- Convert Date to Month ----#
    df_subset['Resource Month'] = pd.to_datetime(df_subset['Resource Date'], dayfirst=True).dt.strftime('%Y-%m')

    # Define a function to calculate actual booked/used slots

    def true_booked(row):
        if row['Total Booked'] <= row['Total Slot Count']:
            return row['Total Booked']
        else:
            return row['Total Slot Count']

    def true_supply(row):
        if row['Total Booked'] <= row['Total Slot Count'] and row['Blocked Status'] == 'N':
            return row['Total Slot Count']
        elif row['Total Booked'] > row['Total Slot Count'] and row['Blocked Status'] == 'N':
            return row['Total Slot Count']
        elif row['Total Booked'] <= row['Total Slot Count'] and row['Blocked Status'] == 'Y':
            return row['Total Booked']
        else:
            return row['Total Slot Count']

    # Create 2 new columns to determine the True Booked and True Supply
    df_subset['True Booked'] = df_subset.apply(true_booked, axis=1)
    df_subset['True Supply'] = df_subset.apply(true_supply, axis=1)

    # Sum the True Booked and True Supply while maintaining the other columns as identifiers
    summed_df_subset = \
        df_subset.groupby(['Resource Date', 'Dept of Resource', 'Resource', 'Rank of Resource', 'Specialty Code',
                           'Clinic', 'Room', 'Frequency', 'SessionType', 'Slot Payment Class Type', 'Slot Visit Type',
                           'Blocked Status', 'Resource Month'])[['True Booked', 'True Supply']].sum().reset_index()

    # Create Identifier
    summed_df_subset['Identifier'] = summed_df_subset[COLS_TO_IDENTIFY].apply(lambda x: '_'.join(x.astype(str)), axis=1)
    summed_df_subset['Priority_level'] = summed_df_subset['Frequency'].map(PRIORITY_LIST)
    summed_df_subset_dedup = summed_df_subset.sort_values('Priority_level').drop_duplicates(subset='Identifier',
                                                                                            keep='first')

    # Re-group dept of resource by specialty
    summed_df_subset_dedup['Department'] = summed_df_subset_dedup['Dept of Resource'].apply(regroup_specialty)
    summarised_df = summed_df_subset_dedup.groupby(
        ['Department', 'Resource Month', 'Slot Visit Type', 'Slot Payment Class Type'])[
        ['True Booked', 'True Supply']].sum().reset_index()

    # Calculate the adjusted values for True Booked and True Supply
    summarised_df['Adjusted True Booked'] = summarised_df.apply(
        lambda x: x['True Booked'] * (0.8 if x['Slot Payment Class Type'] == 'GEN' else 1), axis=1)
    summarised_df['Adjusted True Supply'] = summarised_df.apply(
        lambda x: x['True Supply'] * (0.8 if x['Slot Payment Class Type'] == 'GEN' else 1), axis=1)

    # Group by Specialty (Recoded) and sum the adjusted values
    final_df = summarised_df.groupby(['Department', 'Resource Month'])[
        ['Adjusted True Booked', 'Adjusted True Supply']].sum().round().reset_index()
    final_df.rename(columns={'Adjusted True Supply': 'supply', 'Resource Month': 'months', 'Department': 'department',
                             'Adjusted True Booked': 'slot_booked'},
                    inplace=True)
    print(final_df)
    return final_df


def get_demand(input_file_paths):
    # Read the Excel file

    def filter_for_demand(df):
        df = df[df['Visit Type Code'] == "NC"]  # New Case
        df = df[df['Patient Class Code'].isin(["PSUB", "SUB"])]  # Payment Class filtering
        df = df[df['Appointment Movement Count'] == 0]  # Appointment Movement Count
        df = df[df['Appointment Date Reason Description'].isin(
            ["DOCTOR REQUEST", "EARLIEST DATE", "PATIENT REQUEST", "BLOCK MOVE"])]  # Appointment Reason
        df = df[~df['Attending Doctor Rank Code'].isin([None, 'NDR'])]  # Remove Doctor Rank
        return df

    def filter_for_WTA(df):
        df = df[df['Visit Type Code'] == "NC"]  # New Case
        df = df[df['Patient Class Code'].isin(["PSUB", "SUB"])]  # Payment Class filtering
        df = df[df['Appointment Movement Count'] == 0]  # Appointment Movement Count
        df = df[df['Appointment Date Reason Description'].isin(["EARLIEST DATE"])]  # Appointment Reason
        return df

    master_list = []

    for file in input_file_paths:
        df = pd.read_excel(file)
        df_filtered = filter_for_demand(df)
        df_filtered['Department'] = df_filtered['Specialty Code'].apply(regroup_specialty)
        for (year, month, department), temp in df_filtered.groupby(
                ['Appointment Created Year', 'Appointment Created Month', 'Department']):
            temp_WTA = filter_for_WTA(temp)
            m = create_month_year(year, month)
            cnt = len(temp)
            master_list.append({
                'months': m,
                'department': department,
                'demand': cnt,
                'actual_wta_demand': len(temp_WTA),
                'actual_wta': len(temp_WTA[temp_WTA['Appointment Waiting Time (days)'] <= 60]) / len(temp_WTA)
            })
    return pd.DataFrame(master_list)


def is_keys_exist(excel_file, month, department):
    wb = load_workbook(excel_file)
    ws = wb.active
    idx = 2
    month_idx = 1
    department_idx = 2
    for x in range(1, 100):
        if ws.cell(1, x).value == None:
            break
        if ws.cell(1, x).value.lower() == 'months':
            month_idx = x
        if ws.cell(1, x).value.lower() == 'department':
            department_idx = x
    while True:
        if str(ws.cell(idx, month_idx).value).lower() == str(month).lower():
            if str(ws.cell(idx, department_idx).value).lower() == str(department).lower():
                print(f"{month}-{department} found at row {idx}")
                wb.close()
                return idx
        if ws.cell(idx, 1).value == None:
            wb.close()
            return -1
        idx += 1


def update_data(excel_file, new_data: dict):
    idx = is_keys_exist(excel_file, new_data.get('months'), new_data.get('department'))
    if idx == -1:
        append(excel_file, [[new_data.get('months'), new_data.get('department')]])
        idx = is_keys_exist(excel_file, new_data.get('months'), new_data.get('department'))
    update(excel_file, new_data, idx)
    return new_data


def update(excel_file, new_data, col_idx):
    wb = load_workbook(excel_file)
    ws = wb.active
    # Get Index of New Data
    key_idx = {}
    for x in range(1, 10000):
        if ws.cell(1, x).value == None:
            break
        for k in new_data.keys():
            if ws.cell(1, x).value.lower() == k.lower():
                key_idx[k] = x
    for k, v in new_data.items():
        ws.cell(col_idx, key_idx[k]).value = v
        print(f"Updated at {col_idx, key_idx[k]} {k} with {v}")
    wb.save(excel_file)
    wb.close()


def append(excel_file, new_data):
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in new_data:
        ws.append(row)
    wb.save(excel_file)
    wb.close()


def upload_demand(TEST_DATA):
    if len(TEST_DATA) == 0:
        return {}
    new_d = get_demand(TEST_DATA)
    new_ds = new_d.to_dict(orient='records')
    for d in new_ds:
        update_data(config.DB_FILE, d)
    set_notice("Uploading Step completed", 'small')
    return new_ds


def upload_supply(supply_data):
    if len(supply_data) == 0:
        return {}
    output = []
    for file in supply_data:
        new_d = get_supply(file, config.SHORT_DEPARTMENT)[['months', 'department', 'supply', 'slot_booked']]
        new_ds = new_d.to_dict(orient='records')
        for d in new_ds:
            update_data(config.DB_FILE, d)
        print(f"Uploaded {file}")
        output.extend(new_ds)
        set_notice("Uploading Step completed", 'small')
    return new_ds


# Function to calc WTA for a given month
def _calc_WTA(demand: int, supply: int, slots_booked: int, carry_over: int) -> float:
    slots_avail = supply - slots_booked
    if carry_over <= slots_avail:
        return 0.0
    else:
        WTA = (carry_over - slots_avail) / demand * 100
        if WTA > 100:
            return 100.0
        else:
            return WTA


# Function to calc carry over
def calc_carry_over(demand: int, supply: int, slots_booked: int, carry_over: int) -> int:
    return abs((supply - slots_booked) - carry_over - demand)


# Function will take in a df and modify it to add the carry_over column needed for WTA Calc
def modify_df(df):
    carry_over = []
    for i in range(len(df)):
        if i == 0:
            carry_over.append(0)
        else:
            value = calc_carry_over(df.demand[i - 1], df.supply[i - 1], df.slots_booked[i - 1], carry_over[i - 1])

            carry_over.append(value)
    df['carry_over'] = carry_over
    return df


# Function will take in a df and return the WTA for the months except the last month

def calc_WTA(df, first=True):
    """
    :param df: Dataframe that consist of columns: demand, supply, slots_booked, carry_over
    :param first:
    :return:
    """
    WTA_list = []
    for i in range(len(df) - 1):
        WTA = _calc_WTA(df.demand[i], df.supply[i + 1], df.slots_booked[i + 1], df.carry_over[i + 1])
        WTA_list.append(WTA)
    if first:
        return WTA_list[0]
    else:
        return WTA_list


def get_WTA(demand, supply, slots_booked):
    df = pd.DataFrame({'demand': demand, 'supply': supply, 'slots_booked': slots_booked},
                      columns=['demand', 'supply', 'slots_booked'])
    df = modify_df(df)

    return calc_WTA(df)


def predict_upload(ds):
    for d in ds:
        date = pd.to_datetime(f"{d.get('months')}-01") + relativedelta(months=1)
        _predict_upload(predicted_year=date.year, predicted_month=date.month, department=d.get('department'))


def _predict_upload(predicted_year=2024, predicted_month=1, department='ENT'):
    data = get_training_months(predicted_year=predicted_year, predicted_month=predicted_month, department=department)
    PREDICT_LEN = 3
    MINIMUM_TRAIN = 24
    if len(data) <= MINIMUM_TRAIN:
        return []
    predicted_demand, rmse = get_best_prediction(data, predict_len=PREDICT_LEN)
    print(f"Model: RMSE: {rmse}")
    if len(predicted_demand == PREDICT_LEN):
        timeseries = pd.date_range(start=f"{predicted_year}-{predicted_month}-1", periods=PREDICT_LEN, freq='MS')
        output = []
        for idx, ts in enumerate(timeseries):
            output.append({
                'months': ts.strftime("%Y-%m"),
                'Predicted_Demand': int(predicted_demand.iloc[idx]),
                'department': department
            })
        for d in output:
            update_data(config.DB_FILE, d)
        return output
    else:
        return []


def predict_all(time_start, time_end, departments=['ENT', 'GAS', 'OTO']):
    for dept in departments:
        for date in pd.date_range(start=time_start, end=time_end, freq='MS'):
            print(date, dept)
            print(_predict_upload(date.year, date.month, dept))


if __name__ == "__main__":
    config.refresh_configuration()
    refresh_database()
    predict_all(time_start='2022-02-01', time_end='2024-12-01')
